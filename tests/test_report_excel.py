import io
from datetime import datetime, timezone

from openpyxl import load_workbook

from app.services import reporting_service as rs
from app.services.report_excel import build_reports_workbook


def _sample():
    period = rs.Period(datetime(2026, 7, 21, 19, tzinfo=timezone.utc),
                       datetime(2026, 7, 22, 19, tzinfo=timezone.utc))
    rev = rs.RevenueByMethod(by_method={"cash": 3000, "card": 1500}, gross_tiyn=4500,
                             refunds_tiyn=1500, net_tiyn=3000, orders_count=2)
    top = [rs.ProductRow(name="Латте", qty_sold=3, qty_refunded=1, qty_net=2,
                         revenue_tiyn=4500)]
    cats = [rs.CategoryRow(category="Кофе", revenue_tiyn=4500, qty_net=2)]
    margin = rs.CostMargin(revenue_tiyn=4500, cogs_tiyn=1200, margin_tiyn=3300,
                           margin_pct=73.3, refunds_tiyn=1500, net_revenue_tiyn=3000)
    shifts = rs.ShiftsReport(shifts=[], by_cashier=[])
    return period, rev, top, cats, margin, shifts


def test_build_workbook_has_sheets_and_values():
    data = build_reports_workbook(*_sample())
    wb = load_workbook(io.BytesIO(data))
    assert set(wb.sheetnames) >= {"Сводка", "Способы оплаты", "Топ товаров",
                                  "Категории", "Маржа", "Смены"}
    ws = wb["Способы оплаты"]
    # деньги в тенге: 30 (cash 3000 тиын) и 15 (card 1500 тиын)
    values = [c.value for row in ws.iter_rows() for c in row]
    assert 30 in values and 15 in values
