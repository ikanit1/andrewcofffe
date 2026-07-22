import io

from openpyxl import Workbook
from openpyxl.styles import Font

from app.services import reporting_service as rs

METHOD_LABELS = {"cash": "Наличные", "card": "Карта",
                 "kaspi_qr": "Kaspi QR", "kaspi_terminal": "Kaspi (терминал)"}
_BOLD = Font(bold=True)


def _t(tiyn: int) -> float:
    return round(tiyn / 100, 2)


def _header(ws, titles):
    ws.append(titles)
    for cell in ws[ws.max_row]:
        cell.font = _BOLD


def build_reports_workbook(period: rs.Period, rev: rs.RevenueByMethod,
                           top: list[rs.ProductRow], cats: list[rs.CategoryRow],
                           margin: rs.CostMargin, shifts: rs.ShiftsReport) -> bytes:
    wb = Workbook()

    ws = wb.active
    ws.title = "Сводка"
    _header(ws, ["Показатель", "Значение, тг"])
    ws.append(["Период (UTC)", f"{period.start:%Y-%m-%d %H:%M} — {period.end:%Y-%m-%d %H:%M}"])
    ws.append(["Продажи (валовые)", _t(rev.gross_tiyn)])
    ws.append(["Возвраты", _t(rev.refunds_tiyn)])
    ws.append(["Чистая выручка", _t(rev.net_tiyn)])
    ws.append(["Себестоимость (COGS)", _t(margin.cogs_tiyn)])
    ws.append(["Маржа", _t(margin.margin_tiyn)])
    ws.append(["Маржа, %", margin.margin_pct])
    ws.append(["Чеков", rev.orders_count])

    ws = wb.create_sheet("Способы оплаты")
    _header(ws, ["Способ", "Сумма, тг"])
    for method, amount in rev.by_method.items():
        ws.append([METHOD_LABELS.get(method, method), _t(amount)])
    ws.append(["Возвраты", _t(rev.refunds_tiyn)])
    ws.append(["Чистая", _t(rev.net_tiyn)])

    ws = wb.create_sheet("Топ товаров")
    _header(ws, ["Товар", "Продано", "Возвращено", "Чисто", "Выручка, тг"])
    for r in top:
        ws.append([r.name, r.qty_sold, r.qty_refunded, r.qty_net, _t(r.revenue_tiyn)])

    ws = wb.create_sheet("Категории")
    _header(ws, ["Категория", "Выручка, тг", "Продано (чисто)"])
    for c in cats:
        ws.append([c.category, _t(c.revenue_tiyn), c.qty_net])

    ws = wb.create_sheet("Маржа")
    _header(ws, ["Показатель", "Значение, тг"])
    ws.append(["Выручка", _t(margin.revenue_tiyn)])
    ws.append(["Себестоимость", _t(margin.cogs_tiyn)])
    ws.append(["Маржа", _t(margin.margin_tiyn)])
    ws.append(["Маржа, %", margin.margin_pct])
    ws.append(["Возвраты", _t(margin.refunds_tiyn)])
    ws.append(["Чистая выручка", _t(margin.net_revenue_tiyn)])

    ws = wb.create_sheet("Смены")
    _header(ws, ["Смена", "Кассир", "Открыта", "Закрыта", "Чеков",
                 "Выручка, тг", "Маржа, тг"])
    for sh in shifts.shifts:
        ws.append([sh.shift_id, sh.cashier_name,
                   f"{sh.opened_at:%Y-%m-%d %H:%M}",
                   f"{sh.closed_at:%Y-%m-%d %H:%M}" if sh.closed_at else "",
                   sh.orders_count, _t(sh.revenue_tiyn), _t(sh.margin_tiyn)])
    ws.append([])
    _header(ws, ["Кассир", "Смен", "Чеков", "Выручка, тг", "Маржа, тг"])
    for c in shifts.by_cashier:
        ws.append([c.cashier_name, c.shifts_count, c.orders_count,
                   _t(c.revenue_tiyn), _t(c.margin_tiyn)])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
